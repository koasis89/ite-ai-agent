#!/usr/bin/env python3
"""
Template Binding Converter (openpyxl 기반)

Electron xlsx executor가 spawn하여 사용하는 변환 진입점 스크립트.
표준 양식 템플릿(.xlsx)의 스타일(폰트/테두리/채우기/정렬/숫자서식)을 그대로 유지한 채
정규화된 JSON 레코드를 시트별로 바인딩하여 산출 문서를 저장한다.

사용법:
    python convert.py <payload.json>
    echo '<payload json>' | python convert.py -

payload 스키마:
{
  "templatePath": "<원본 .xlsx 절대경로>",  # 생략/빈 값이면 템플릿 없이 새 워크북 생성(스크래치 모드)
  "outputPath":   "<저장 .xlsx 절대경로>",
  "metadataSheets": ["문서정보"],          # 데이터로 덮어쓰면 안 되는 메타 시트명(선택)
  "recalc": false,                          # true면 저장 후 recalc.py로 수식 재계산(선택)
  "sheets": [
    {
      "name": "18.화면서비스호출",
      "dataStartRow": 2,
      "headers": ["화면 ID", "화면명", ...],  # 스크래치 모드에서 1행에 기록할 헤더(선택)
      "columns": [
        {"idx": 1, "field": "screenId", "type": "string"},
        ...
      ],
      "records": [ {"screenId": "...", ...}, ... ]
    }
  ]
}

결과(JSON, stdout):
{
  "status": "success" | "error",
  "outputPath": "...",
  "sheetsWritten": [{"requested": "...", "applied": "...", "rows": N}],
  "error": "..."   # 실패 시
}
"""

import json
import sys
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _coerce(value, col_type):
    """셀 값을 지정된 타입으로 강제 변환한다."""
    if value is None:
        return "" if col_type == "string" else (0 if col_type == "number" else False)

    if col_type == "number":
        if isinstance(value, (int, float)):
            return value
        try:
            cleaned = "".join(ch for ch in str(value) if ch.isdigit() or ch in ".-")
            return float(cleaned) if cleaned not in ("", "-", ".") else 0
        except (ValueError, TypeError):
            return 0

    if col_type == "boolean":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("true", "1", "yes", "y", "완료", "참")

    return str(value)


def _copy_style(src_cell, dst_cell):
    """openpyxl 셀 스타일을 깊은 복사로 이식한다."""
    if src_cell is None or not getattr(src_cell, "has_style", False):
        return
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)


def _resolve_worksheet(workbook, sheet_name, metadata_sheets):
    """시트명으로 워크시트를 찾고, 없으면 메타 시트를 건너뛴 첫 데이터 시트로 대피한다."""
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name], sheet_name

    for ws in workbook.worksheets:
        if ws.title not in metadata_sheets:
            return ws, ws.title

    return workbook.worksheets[0], workbook.worksheets[0].title


def _bind_sheet(workbook, sheet_cfg, metadata_sheets):
    name = sheet_cfg["name"]
    data_start_row = int(sheet_cfg["dataStartRow"])
    columns = sheet_cfg.get("columns", [])
    records = sheet_cfg.get("records", [])

    worksheet, applied = _resolve_worksheet(workbook, name, metadata_sheets)

    max_col = max((int(c["idx"]) for c in columns), default=1)

    # 1. 데이터 시작 행의 원본 스타일을 백업(복사 레퍼런스)
    source_styles = {}
    for col_idx in range(1, max_col + 1):
        source_styles[col_idx] = worksheet.cell(row=data_start_row, column=col_idx)
    source_style_snapshot = {
        col_idx: (copy(cell._style) if cell.has_style else None)
        for col_idx, cell in source_styles.items()
    }
    source_height = worksheet.row_dimensions[data_start_row].height

    # 2. 레코드가 2개 이상이면 하단으로 밀어내며 행 삽입
    if len(records) > 1:
        worksheet.insert_rows(data_start_row + 1, amount=len(records) - 1)

    # 3. 값 주입 + 스타일 오버레이
    for i, record in enumerate(records):
        row_idx = data_start_row + i
        for col in columns:
            col_idx = int(col["idx"])
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = _coerce(record.get(col["field"]), col.get("type", "string"))
            snapshot = source_style_snapshot.get(col_idx)
            if snapshot is not None:
                cell._style = copy(snapshot)
        if source_height:
            worksheet.row_dimensions[row_idx].height = source_height

    return {"requested": name, "applied": applied, "rows": len(records)}


def _bind_sheet_scratch(worksheet, sheet_cfg):
    """템플릿이 없는 새 워크북에 헤더/데이터를 직접 기록한다."""
    data_start_row = int(sheet_cfg["dataStartRow"])
    columns = sheet_cfg.get("columns", [])
    records = sheet_cfg.get("records", [])
    headers = sheet_cfg.get("headers", [])

    if headers:
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=str(header))

    for i, record in enumerate(records):
        row_idx = data_start_row + i
        for col in columns:
            col_idx = int(col["idx"])
            worksheet.cell(row=row_idx, column=col_idx).value = _coerce(
                record.get(col["field"]), col.get("type", "string")
            )

    return {"requested": sheet_cfg["name"], "applied": worksheet.title, "rows": len(records)}


def _maybe_recalc(output_path):
    """recalc.py를 통해 수식 재계산을 시도한다(LibreOffice 필요). 실패해도 변환 자체는 성공으로 둔다."""
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from recalc import recalc  # noqa: E402

        return recalc(output_path)
    except Exception as exc:  # noqa: BLE001
        return {"status": "recalc_skipped", "reason": str(exc)}


def convert(payload):
    template_path = payload.get("templatePath")
    output_path = payload["outputPath"]
    metadata_sheets = set(payload.get("metadataSheets", ["문서정보"]))
    sheets = payload.get("sheets", [])

    sheets_written = []

    if template_path:
        # 템플릿 모드: 지정된 템플릿을 상속해 스타일을 유지한 채 바인딩
        if not Path(template_path).exists():
            return {"status": "error", "error": f"템플릿 파일을 찾을 수 없습니다: {template_path}"}
        workbook = load_workbook(template_path)
        for sheet_cfg in sheets:
            sheets_written.append(_bind_sheet(workbook, sheet_cfg, metadata_sheets))
    else:
        # 스크래치 모드: 템플릿 없이 새 워크북을 생성해 직접 기록
        workbook = Workbook()
        default_ws = workbook.active
        for sheet_idx, sheet_cfg in enumerate(sheets):
            title = str(sheet_cfg.get("name", f"Sheet{sheet_idx + 1}"))[:31]
            if sheet_idx == 0:
                worksheet = default_ws
                worksheet.title = title
            else:
                worksheet = workbook.create_sheet(title=title)
            sheets_written.append(_bind_sheet_scratch(worksheet, sheet_cfg))
        if not sheets:
            default_ws.title = "Sheet1"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()

    result = {
        "status": "success",
        "outputPath": output_path,
        "sheetsWritten": sheets_written,
    }

    if payload.get("recalc"):
        result["recalc"] = _maybe_recalc(output_path)

    return result


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "error": "payload 인자가 필요합니다."}))
        sys.exit(1)

    arg = sys.argv[1]
    try:
        raw = sys.stdin.read() if arg == "-" else Path(arg).read_text(encoding="utf-8-sig")
        payload = json.loads(raw)
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"status": "error", "error": f"payload 파싱 실패: {exc}"}))
        sys.exit(1)

    try:
        result = convert(payload)
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False))
        sys.exit(1)

    print(json.dumps(result, ensure_ascii=False))
    sys.exit(0 if result.get("status") == "success" else 1)


if __name__ == "__main__":
    main()
